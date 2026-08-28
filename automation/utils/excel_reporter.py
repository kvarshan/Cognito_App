import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelReporter:
    @staticmethod
    def generate_excel_reports(test_results, output_dir):
        os.makedirs(output_dir, exist_ok=True)
        
        # 1. Generate Full Automation Test Report
        full_report_path = os.path.join(output_dir, "Automation_Test_Report.xlsx")
        wb = openpyxl.Workbook()
        
        # Header Style
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        skip_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        
        pass_font = Font(name="Calibri", size=10, bold=True, color="166534")
        fail_font = Font(name="Calibri", size=10, bold=True, color="991B1B")
        skip_font = Font(name="Calibri", size=10, bold=True, color="92400E")
        
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # ----------------------------------------------------
        # Sheet 1: Executed Test Cases
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        headers1 = ["Test ID", "Suite / Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Expected Result", "Actual Result"]
        ws1.append(headers1)
        
        for col_num, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, test in enumerate(test_results, 2):
            status = test.get("status", "PASS")
            ws1.append([
                test.get("id", f"TC_{row_idx-1:04d}"),
                test.get("module", "General"),
                test.get("name", "Test Execution"),
                status,
                test.get("duration", 0.05),
                test.get("priority", "P1"),
                test.get("expected", "Success"),
                test.get("actual", "Success")
            ])
            status_cell = ws1.cell(row=row_idx, column=4)
            if status == "PASS":
                status_cell.fill = pass_fill
                status_cell.font = pass_font
            elif status == "FAIL":
                status_cell.fill = fail_fill
                status_cell.font = fail_font
            else:
                status_cell.fill = skip_fill
                status_cell.font = skip_font

        # ----------------------------------------------------
        # Sheet 2: Passed Tests
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Passed Tests")
        ws2.append(headers1)
        for col_num in range(1, len(headers1)+1):
            ws2.cell(row=1, column=col_num).font = header_font
            ws2.cell(row=1, column=col_num).fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
            
        passed_tests = [t for t in test_results if t.get("status") == "PASS"]
        for t in passed_tests:
            ws2.append([t.get("id"), t.get("module"), t.get("name"), "PASS", t.get("duration", 0.05), t.get("priority"), t.get("expected"), t.get("actual")])

        # ----------------------------------------------------
        # Sheet 3: Failed Tests
        # ----------------------------------------------------
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.append(headers1)
        for col_num in range(1, len(headers1)+1):
            ws3.cell(row=1, column=col_num).font = header_font
            ws3.cell(row=1, column=col_num).fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
        failed_tests = [t for t in test_results if t.get("status") == "FAIL"]
        for t in failed_tests:
            ws3.append([t.get("id"), t.get("module"), t.get("name"), "FAIL", t.get("duration", 0.05), t.get("priority"), t.get("expected"), t.get("actual")])

        # ----------------------------------------------------
        # Sheet 4: Skipped Tests
        # ----------------------------------------------------
        ws4 = wb.create_sheet(title="Skipped Tests")
        ws4.append(headers1)
        for col_num in range(1, len(headers1)+1):
            ws4.cell(row=1, column=col_num).font = header_font
            ws4.cell(row=1, column=col_num).fill = PatternFill(start_color="92400E", end_color="92400E", fill_type="solid")
        skipped_tests = [t for t in test_results if t.get("status") in ("SKIP", "SKIPPED")]
        for t in skipped_tests:
            ws4.append([t.get("id"), t.get("module"), t.get("name"), "SKIPPED", t.get("duration", 0.05), t.get("priority"), t.get("expected"), t.get("actual")])

        # ----------------------------------------------------
        # Sheet 5: Execution Metrics
        # ----------------------------------------------------
        ws5 = wb.create_sheet(title="Execution Metrics")
        ws5.append(["Metric", "Value"])
        ws5.cell(row=1, column=1).font = header_font
        ws5.cell(row=1, column=1).fill = header_fill
        ws5.cell(row=1, column=2).font = header_font
        ws5.cell(row=1, column=2).fill = header_fill
        
        total_count = len(test_results)
        passed_count = len(passed_tests)
        failed_count = len(failed_tests)
        skipped_count = len(skipped_tests)
        pass_rate = f"{(passed_count / total_count * 100):.2f}%" if total_count > 0 else "0%"
        total_dur = sum(t.get("duration", 0.05) for t in test_results)
        
        metrics = [
            ("Total Test Cases", total_count),
            ("Passed Test Cases", passed_count),
            ("Failed Test Cases", failed_count),
            ("Skipped Test Cases", skipped_count),
            ("Overall Pass Percentage", pass_rate),
            ("Total Execution Time (s)", f"{total_dur:.2f}s"),
            ("Target Deployment URL", "https://kvarshan.github.io/Cognito_App/"),
            ("Pipeline Status", "PASSED" if failed_count == 0 else "FAILED")
        ]
        for row in metrics:
            ws5.append(list(row))

        # ----------------------------------------------------
        # Sheet 6: Defect Summary
        # ----------------------------------------------------
        ws6 = wb.create_sheet(title="Defect Summary")
        ws6.append(["Defect ID", "Test Case ID", "Module", "Severity", "Description", "Status"])
        for col_num in range(1, 7):
            ws6.cell(row=1, column=col_num).font = header_font
            ws6.cell(row=1, column=col_num).fill = header_fill
            
        if not failed_tests:
            ws6.append(["N/A", "N/A", "All Suites", "None", "Zero critical defects found. All test cases passed.", "RESOLVED"])

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    except:
                        pass
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(full_report_path)
        
        # Also create standalone auxiliary workbooks
        ExcelReporter._save_subset_workbook(passed_tests, headers1, header_font, "166534", os.path.join(output_dir, "Passed_Test_Cases.xlsx"))
        ExcelReporter._save_subset_workbook(failed_tests, headers1, header_font, "991B1B", os.path.join(output_dir, "Failed_Test_Cases.xlsx"))
        ExcelReporter._save_summary_workbook(metrics, header_font, header_fill, os.path.join(output_dir, "Summary_Report.xlsx"))

        return full_report_path

    @staticmethod
    def _save_subset_workbook(tests, headers, header_font, color_hex, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        for t in tests:
            ws.append([t.get("id"), t.get("module"), t.get("name"), t.get("status"), t.get("duration", 0.05), t.get("priority"), t.get("expected"), t.get("actual")])
        wb.save(file_path)

    @staticmethod
    def _save_summary_workbook(metrics, header_font, header_fill, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        ws.append(["Key Performance Metric", "Result Value"])
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2).font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        for m in metrics:
            ws.append(list(m))
        wb.save(file_path)
